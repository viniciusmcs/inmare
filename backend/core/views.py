import os
import tempfile
import hashlib
import mimetypes
import secrets
from io import BytesIO
from xml.sax.saxutils import escape
from django.conf import settings as django_settings
from django.contrib.auth import get_user_model
from django.db import connection, transaction
from django.db.models import Case, Count, IntegerField, Max, Min, Q, Sum, Value, When
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from datetime import datetime, time, timedelta
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.text import slugify
from rest_framework import mixins, permissions, status, viewsets
from rest_framework.pagination import PageNumberPagination
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter
from rest_framework.response import Response
from rest_framework.throttling import AnonRateThrottle
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from .models import AuditEvent, Broker, CRMActivity, CRMContact, CRMImportBatch, CRMImportRow, CRMNotification, CRMOpportunity, CRMProposal, CRMPropertyLink, CRMTask, Development, FrequentlyAskedQuestion, HeroSlide, ImportJob, InstitutionalImage, Lead, ListingOption, Media, Property, SiteSettings, Testimonial, catalog_key, person_name_search_key
from .serializers import AdminLeadSerializer, AdminUserSerializer, AuditSerializer, AdminPropertySerializer, BrokerSerializer, CRMActivitySerializer, CRMContactChoiceSerializer, CRMContactPoolSerializer, CRMContactSerializer, CRMImportBatchSerializer, CRMImportRowSerializer, CRMNotificationSerializer, CRMOpportunitySerializer, CRMProposalSerializer, CRMPropertyLinkSerializer, CRMTaskSerializer, DevelopmentSerializer, FrequentlyAskedQuestionSerializer, HeroSlideSerializer, ImportJobSerializer, InstitutionalImageSerializer, LeadSerializer, ListingOptionSerializer, PublicPropertySerializer, SiteSettingsSerializer, TestimonialSerializer
from .services import extract_property_description, import_property_folder, import_property_zip
from .crm_services import commit_import_batch, file_sha256, find_duplicate_contact, find_duplicate_import_row, process_import_batch, sanitize_import_row, validate_crm_import
from .crm_permissions import CanManageProperties, IsCRMManager, IsCRMUser, can_view_all_crm, crm_user_payload, user_broker
from .media_utils import normalize_uploaded_image

class LeadThrottle(AnonRateThrottle): scope = "lead"
class LoginThrottle(AnonRateThrottle): scope = "login"
class PublicPropertyPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 20


class CRMContactPagination(PageNumberPagination):
    page_size = 30
    page_size_query_param = "page_size"
    max_page_size = 50

class HealthView(APIView):
    permission_classes = [permissions.AllowAny]
    def get(self, request):
        try:
            with connection.cursor() as cursor: cursor.execute("SELECT 1")
            return Response({"status": "ok", "database": "ok"})
        except Exception:
            return Response({"status": "error", "database": "error"}, status=503)

def absolute_public_url(request, path):
    configured = os.getenv("SITE_URL", "").rstrip("/")
    base = configured or request.build_absolute_uri("/").rstrip("/")
    return f"{base}{path}"

def robots_txt(request):
    lines = [
        "User-agent: *",
        "Allow: /",
        "Disallow: /admin/",
        "Disallow: /django-admin/",
        "Disallow: /api/v1/admin/",
        f"Sitemap: {absolute_public_url(request, '/sitemap.xml')}",
        "",
    ]
    return HttpResponse("\n".join(lines), content_type="text/plain; charset=utf-8")

def sitemap_xml(request):
    static_paths = ["/", "/imoveis", "/imobiliaria", "/empreendimentos", "/encontrar-imovel", "/anuncie-seu-imovel", "/contato"]
    urls = [(path, timezone.now(), "weekly", "0.8") for path in static_paths]
    properties = Property.objects.filter(
        published=True,
        hidden=False,
        archived_at__isnull=True,
        status__in=[Property.Status.AVAILABLE, Property.Status.RESERVED, Property.Status.NEGOTIATING],
    ).only("slug", "updated_at")
    developments = Development.objects.filter(published=True, hidden=False).only("slug", "updated_at")
    urls += [(f"/imoveis/{item.slug}", item.updated_at, "daily", "0.9") for item in properties]
    urls += [(f"/empreendimentos/{item.slug}", item.updated_at, "weekly", "0.7") for item in developments]
    body = ['<?xml version="1.0" encoding="UTF-8"?>', '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for path, updated_at, changefreq, priority in urls:
        body.extend([
            "  <url>",
            f"    <loc>{escape(absolute_public_url(request, path))}</loc>",
            f"    <lastmod>{updated_at.date().isoformat()}</lastmod>",
            f"    <changefreq>{changefreq}</changefreq>",
            f"    <priority>{priority}</priority>",
            "  </url>",
        ])
    body.append("</urlset>")
    return HttpResponse("\n".join(body), content_type="application/xml; charset=utf-8")

class LoginView(APIView):
    permission_classes = [permissions.AllowAny]
    authentication_classes = []
    throttle_classes = [] if django_settings.DEBUG else [LoginThrottle]
    def post(self, request):
        from django.contrib.auth import authenticate
        user = authenticate(username=request.data.get("username"), password=request.data.get("password"))
        if not user or not (user.is_staff or user_broker(user)): return Response({"detail": "Credenciais inválidas."}, status=401)
        refresh = RefreshToken.for_user(user)
        response = Response({"user": crm_user_payload(user)})
        response.set_cookie("access_token", str(refresh.access_token), httponly=True, secure=not django_settings.DEBUG, samesite="Lax")
        response.set_cookie("refresh_token", str(refresh), httponly=True, secure=not django_settings.DEBUG, samesite="Lax")
        AuditEvent.objects.create(actor=user, action="auth.login", entity_type="User", entity_id=str(user.id))
        return response


class CurrentUserView(APIView):
    permission_classes = [IsCRMUser]

    def get(self, request):
        return Response(crm_user_payload(request.user))

class LogoutView(APIView):
    def post(self, request):
        raw_refresh = request.COOKIES.get("refresh_token")
        if raw_refresh:
            try: RefreshToken(raw_refresh).blacklist()
            except Exception: pass
        response = Response(status=204)
        response.delete_cookie("access_token"); response.delete_cookie("refresh_token")
        AuditEvent.objects.create(actor=request.user, action="auth.logout", entity_type="User", entity_id=str(request.user.id))
        return response

class RefreshCookieView(APIView):
    permission_classes = [permissions.AllowAny]
    authentication_classes = []
    def post(self, request):
        raw_refresh = request.COOKIES.get("refresh_token")
        if not raw_refresh: return Response({"detail": "Refresh token ausente."}, status=401)
        try:
            refresh = RefreshToken(raw_refresh)
            response = Response({"detail": "Token renovado."})
            response.set_cookie("access_token", str(refresh.access_token), httponly=True, secure=not django_settings.DEBUG, samesite="Lax")
            return response
        except Exception:
            return Response({"detail": "Refresh token inválido."}, status=401)

class PublicPropertyViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [permissions.AllowAny]
    authentication_classes = []
    serializer_class = PublicPropertySerializer
    lookup_field = "slug"
    pagination_class = PublicPropertyPagination
    filterset_fields = ("public_id", "property_type", "purpose", "city", "neighborhood", "bedrooms", "suites", "featured", "launch", "exclusive")
    search_fields = ("title", "public_description", "city", "neighborhood")
    ordering_fields = ("price", "created_at", "title", "private_area", "bedrooms", "suites", "featured")
    def get_queryset(self):
        queryset = Property.objects.filter(
            published=True,
            hidden=False,
            archived_at__isnull=True,
            status__in=[
                Property.Status.AVAILABLE,
                Property.Status.RESERVED,
                Property.Status.NEGOTIATING,
            ],
        ).prefetch_related("media").order_by("-featured", "-created_at")
        if self.request.query_params.get("launches") == "true":
            queryset = queryset.filter(
                Q(launch=True) | Q(created_at__gte=timezone.now() - timedelta(days=7))
            ).order_by("-launch", "-created_at")
        max_price = self.request.query_params.get("max_price")
        min_price = self.request.query_params.get("min_price")
        min_area = self.request.query_params.get("min_area")
        max_area = self.request.query_params.get("max_area")
        bathrooms = self.request.query_params.get("bathrooms")
        parking_spaces = self.request.query_params.get("parking_spaces")
        features = self.request.query_params.getlist("feature")
        if max_price:
            queryset = queryset.filter(price__lte=max_price)
        if min_price:
            queryset = queryset.filter(price__gte=min_price)
        if min_area:
            queryset = queryset.filter(private_area__gte=min_area)
        if max_area:
            queryset = queryset.filter(private_area__lte=max_area)
        if bathrooms:
            queryset = queryset.filter(bathrooms__gte=bathrooms)
        if parking_spaces:
            queryset = queryset.filter(parking_spaces__gte=parking_spaces)
        if self.request.query_params.get("has_video") == "true":
            queryset = queryset.filter(media__kind=Media.Kind.VIDEO, media__status=Media.Status.READY)
        for feature in features:
            queryset = queryset.filter(features__icontains=feature)
        if self.request.query_params.get("accepts_financing") == "true":
            queryset = queryset.filter(accepts_financing=True)
        if self.request.query_params.get("accepts_exchange") == "true":
            queryset = queryset.filter(accepts_exchange=True)
        return queryset.distinct()

class PublicDevelopmentViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [permissions.AllowAny]
    authentication_classes = []
    serializer_class = DevelopmentSerializer
    lookup_field = "slug"
    def get_queryset(self): return Development.objects.filter(published=True, hidden=False).prefetch_related("media")

class PublicSettingsView(APIView):
    permission_classes = [permissions.AllowAny]
    authentication_classes = []
    def get(self, request):
        settings = SiteSettings.objects.first()
        return Response(SiteSettingsSerializer(settings).data if settings else {})

class PublicContentView(APIView):
    permission_classes = [permissions.AllowAny]
    authentication_classes = []
    def get(self, request):
        return Response({
            "hero_slides": HeroSlideSerializer(HeroSlide.objects.filter(active=True)[:12], many=True).data,
            "testimonials": TestimonialSerializer(Testimonial.objects.filter(active=True), many=True).data,
            "faqs": FrequentlyAskedQuestionSerializer(FrequentlyAskedQuestion.objects.filter(active=True), many=True).data,
            "institutional_images": InstitutionalImageSerializer(InstitutionalImage.objects.filter(active=True), many=True).data,
        })

class PublicFilterOptionsView(APIView):
    permission_classes = [permissions.AllowAny]
    authentication_classes = []
    def get(self, request):
        queryset = Property.objects.filter(
            published=True, hidden=False, archived_at__isnull=True,
            status__in=[Property.Status.AVAILABLE, Property.Status.RESERVED, Property.Status.NEGOTIATING],
        )
        def values(field, text=False):
            result = queryset.exclude(**{f"{field}__isnull": True})
            if text:
                result = result.exclude(**{field: ""})
            return list(result.values_list(field, flat=True).distinct().order_by(field))
        return Response({
            "property_types": values("property_type", True),
            "cities": values("city", True),
            "neighborhoods": values("neighborhood", True),
            "bedrooms": values("bedrooms"),
            "suites": values("suites"),
            "bathrooms": values("bathrooms"),
            "parking_spaces": values("parking_spaces"),
            "features": sorted({feature for items in queryset.values_list("features", flat=True) for feature in (items or [])}),
            "min_price": queryset.aggregate(value=Min("price"))["value"] or 0,
            "max_price": queryset.aggregate(value=Max("price"))["value"] or 0,
            "max_area": queryset.aggregate(value=Max("private_area"))["value"] or 0,
        })

class LeadViewSet(mixins.CreateModelMixin, viewsets.GenericViewSet):
    permission_classes = [permissions.AllowAny]
    authentication_classes = []
    throttle_classes = [LeadThrottle]
    serializer_class = LeadSerializer
    queryset = Lead.objects.all()


class WhatsAppPropertyIngestView(APIView):
    """Receives normalized WhatsApp events from the private n8n workflow."""

    permission_classes = [permissions.AllowAny]
    authentication_classes = []

    def post(self, request):
        expected_token = django_settings.WHATSAPP_INGEST_TOKEN
        supplied_token = request.headers.get("X-Inmare-Ingest-Token", "")
        if not expected_token:
            return Response({"detail": "Integração não configurada."}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        if not supplied_token or not secrets.compare_digest(supplied_token, expected_token):
            return Response({"detail": "Credencial de integração inválida."}, status=status.HTTP_401_UNAUTHORIZED)

        if request.data.get("accepted") is not True:
            return Response({"accepted": False, "detail": "Evento ignorado pelo filtro do fluxo."}, status=status.HTTP_202_ACCEPTED)

        normalized = request.data.get("normalized")
        batch = request.data.get("batch")
        if not isinstance(normalized, dict) or not isinstance(batch, dict):
            raise ValidationError({"detail": "Evento normalizado e lote são obrigatórios."})

        group_id = str(normalized.get("remoteJid") or "").strip()
        configured_group_id = django_settings.WHATSAPP_INGEST_GROUP_ID.strip()
        if configured_group_id and group_id != configured_group_id:
            return Response({"detail": "Grupo não autorizado."}, status=status.HTTP_403_FORBIDDEN)

        batch_id = str(batch.get("id") or "").strip()
        message_id = str(normalized.get("messageId") or "").strip()
        if not batch_id or not message_id:
            raise ValidationError({"detail": "Identificadores do lote e da mensagem são obrigatórios."})

        external_id = f"whatsapp:{batch_id}"[:160]
        duplicate = AuditEvent.objects.filter(
            action="property.whatsapp_message_received",
            metadata__message_id=message_id,
            metadata__group_id=group_id,
        ).first()
        if duplicate:
            prop = Property.objects.filter(pk=duplicate.entity_id).first()
            if prop:
                return Response(self._response(prop, created=False, duplicate=True))

        text = str(normalized.get("text") or "").replace("\x00", "").strip()[:12000]
        media_type = str(normalized.get("mediaType") or "").strip()[:30]
        file_name = str(normalized.get("fileName") or "").replace("\x00", "").strip()[:240]
        sender = str(normalized.get("sender") or "").strip()[:120]
        push_name = str(normalized.get("pushName") or "").replace("\x00", "").strip()[:160]

        with transaction.atomic():
            prop = Property.objects.select_for_update().filter(
                source="whatsapp",
                external_id=external_id,
            ).first()
            is_start_marker = bool(batch.get("startMarkerDetected"))
            if prop is None and not is_start_marker:
                recent_event = AuditEvent.objects.filter(
                    action="property.whatsapp_message_received",
                    metadata__group_id=group_id,
                    created_at__gte=timezone.now() - timedelta(minutes=10),
                ).order_by("-created_at").first()
                if recent_event:
                    prop = Property.objects.select_for_update().filter(
                        pk=recent_event.entity_id,
                        source="whatsapp",
                        status=Property.Status.DRAFT,
                        published=False,
                    ).first()
            created = prop is None
            if created:
                placeholder = f"[WhatsApp] Imóvel recebido em {timezone.localtime():%d/%m/%Y %H:%M}"
                prop = Property.objects.create(
                    title=placeholder,
                    slug=self._unique_slug(placeholder, external_id),
                    public_description="",
                    property_type="",
                    purpose=Property.Purpose.SALE,
                    status=Property.Status.DRAFT,
                    city="",
                    neighborhood="",
                    published=False,
                    hidden=False,
                    featured=False,
                    source="whatsapp",
                    external_id=external_id,
                    internal_notes=(
                        "Recebido automaticamente do WhatsApp.\n"
                        "Status: aguardando revisão administrativa.\n"
                        f"Grupo: {group_id}\n"
                        f"Lote: {batch_id}\n"
                    ),
                )

            entry = [f"\n--- Mensagem {message_id} ---"]
            if push_name or sender:
                entry.append(f"Remetente: {push_name or 'Não informado'} ({sender or 'ID não informado'})")
            if text:
                entry.append(text)
            if media_type:
                media_label = f"Mídia recebida: {media_type}"
                if file_name:
                    media_label += f" — {file_name}"
                entry.append(media_label)
            prop.internal_notes = (prop.internal_notes + "\n".join(entry))[-500000:]
            self._apply_safe_suggestions(prop, text, is_start_marker)
            prop.save()

            AuditEvent.objects.create(
                action="property.whatsapp_message_received",
                entity_type="Property",
                entity_id=str(prop.id),
                metadata={
                    "automatic": True,
                    "batch_id": batch_id,
                    "group_id": group_id,
                    "message_id": message_id,
                    "media_type": media_type,
                    "created_property": created,
                },
            )

        return Response(
            self._response(prop, created=created, duplicate=False),
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )

    @staticmethod
    def _unique_slug(title, external_id):
        suffix = hashlib.sha256(external_id.encode("utf-8")).hexdigest()[:8]
        base = (slugify(title) or "imovel-whatsapp")[:205]
        candidate = f"{base}-{suffix}"[:220]
        counter = 2
        while Property.objects.filter(slug=candidate).exists():
            candidate = f"{base[:200]}-{suffix}-{counter}"[:220]
            counter += 1
        return candidate

    @staticmethod
    def _catalog_value(kind, value, city=""):
        if not value:
            return ""
        filters = {"kind": kind, "key": catalog_key(str(value)), "active": True}
        if kind == ListingOption.Kind.NEIGHBORHOOD:
            filters["city_key"] = catalog_key(city)
        option = ListingOption.objects.filter(**filters).first()
        return option.name if option else ""

    def _apply_safe_suggestions(self, prop, text, is_start_marker):
        if not text:
            return
        suggestions = extract_property_description(text)

        def value(field):
            suggestion = suggestions.get(field)
            return suggestion.get("value") if isinstance(suggestion, dict) else None

        suggested_title = value("title")
        if suggested_title and not is_start_marker and prop.title.startswith("[WhatsApp]"):
            prop.title = str(suggested_title)[:200]
            prop.slug = self._unique_slug(prop.title, prop.external_id)

        for field in (
            "price", "condominium_fee", "iptu", "bedrooms", "suites", "bathrooms",
            "parking_spaces", "private_area", "total_area", "land_dimensions",
            "solar_orientation", "accepts_financing", "accepts_exchange",
            "private_address", "private_commission",
        ):
            suggested = value(field)
            if suggested not in (None, "", []) and getattr(prop, field) in (None, "", []):
                setattr(prop, field, suggested)

        suggested_features = value("features")
        if isinstance(suggested_features, list):
            prop.features = list(dict.fromkeys([*(prop.features or []), *suggested_features]))[:100]

        if not prop.property_type:
            prop.property_type = self._catalog_value(ListingOption.Kind.PROPERTY_TYPE, value("property_type"))
        if not prop.city:
            prop.city = self._catalog_value(ListingOption.Kind.CITY, value("city"))
        if not prop.neighborhood and prop.city:
            prop.neighborhood = self._catalog_value(
                ListingOption.Kind.NEIGHBORHOOD,
                value("neighborhood"),
                city=prop.city,
            )

    @staticmethod
    def _response(prop, *, created, duplicate):
        return {
            "accepted": True,
            "created": created,
            "duplicate": duplicate,
            "property_id": str(prop.id),
            "public_id": prop.public_id,
            "status": prop.status,
            "published": prop.published,
            "source": prop.source,
            "review_status": "pending_admin_review" if not prop.published else "published",
        }


class AdminListingOptionViewSet(
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    permission_classes = [permissions.IsAdminUser]
    serializer_class = ListingOptionSerializer
    pagination_class = None

    def get_permissions(self):
        if self.action == "list":
            return [CanManageProperties()]
        return [permissions.IsAdminUser()]

    def get_queryset(self):
        return ListingOption.objects.filter(active=True)

    def perform_create(self, serializer):
        option = serializer.save()
        AuditEvent.objects.create(
            actor=self.request.user,
            action="listing_option.created",
            entity_type="ListingOption",
            entity_id=str(option.id),
            metadata={"kind": option.kind, "name": option.name, "city": option.city},
        )

    def perform_update(self, serializer):
        with transaction.atomic():
            current = ListingOption.objects.select_for_update().get(pk=serializer.instance.pk)
            old_name = current.name
            old_city = current.city
            serializer.instance = current
            option = serializer.save()
            if option.kind == ListingOption.Kind.PROPERTY_TYPE:
                updated = Property.objects.filter(property_type=old_name).update(property_type=option.name)
            elif option.kind == ListingOption.Kind.CITY:
                updated = Property.objects.filter(city=old_name).update(city=option.name)
                ListingOption.objects.filter(
                    kind=ListingOption.Kind.NEIGHBORHOOD,
                    city=old_name,
                ).update(city=option.name, city_key=catalog_key(option.name))
            else:
                updated = Property.objects.filter(
                    city=old_city,
                    neighborhood=old_name,
                ).update(neighborhood=option.name)
            AuditEvent.objects.create(
                actor=self.request.user,
                action="listing_option.updated",
                entity_type="ListingOption",
                entity_id=str(option.id),
                metadata={
                    "kind": option.kind,
                    "old_name": old_name,
                    "new_name": option.name,
                    "city": option.city,
                    "properties_updated": updated,
                },
            )

    def destroy(self, request, *args, **kwargs):
        option = self.get_object()
        if option.kind == ListingOption.Kind.PROPERTY_TYPE:
            properties_in_use = Property.objects.filter(property_type__iexact=option.name).count()
            neighborhoods_in_use = 0
        elif option.kind == ListingOption.Kind.CITY:
            properties_in_use = Property.objects.filter(city__iexact=option.name).count()
            neighborhoods_in_use = ListingOption.objects.filter(
                kind=ListingOption.Kind.NEIGHBORHOOD,
                city__iexact=option.name,
            ).count()
        else:
            properties_in_use = Property.objects.filter(
                city__iexact=option.city,
                neighborhood__iexact=option.name,
            ).count()
            neighborhoods_in_use = 0
        blockers = []
        if properties_in_use:
            blockers.append(f"{properties_in_use} imóvel(is)")
        if neighborhoods_in_use:
            blockers.append(f"{neighborhoods_in_use} bairro(s)")
        if blockers:
            raise ValidationError({
                "detail": (
                    f'Não é possível excluir “{option.name}” porque está vinculado a '
                    f'{" e ".join(blockers)}. Altere os vínculos antes de excluir.'
                )
            })
        option_id = str(option.id)
        metadata = {"kind": option.kind, "name": option.name, "city": option.city}
        with transaction.atomic():
            option.delete()
            AuditEvent.objects.create(
                actor=request.user,
                action="listing_option.deleted",
                entity_type="ListingOption",
                entity_id=option_id,
                metadata=metadata,
            )
        return Response(status=status.HTTP_204_NO_CONTENT)


class AdminPropertyViewSet(viewsets.ModelViewSet):
    permission_classes = [CanManageProperties]
    serializer_class = AdminPropertySerializer
    queryset = Property.objects.all().select_related("broker").prefetch_related("media")
    filterset_fields = ("status", "published", "hidden", "featured", "launch", "city")
    search_fields = ("title", "public_id", "city", "neighborhood")

    ADMIN_ONLY_ACTIONS = {
        "destroy", "publish", "validate_media", "toggle_featured", "toggle_launch",
        "mark_in_service", "remove_in_service", "mark_sold", "restore_sale",
        "confirm_review", "archive", "restore_archive",
    }

    def get_permissions(self):
        if self.action in self.ADMIN_ONLY_ACTIONS:
            return [permissions.IsAdminUser()]
        return [CanManageProperties()]

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_staff:
            queryset = queryset.filter(broker=user_broker(self.request.user))
        return queryset.annotate(
            whatsapp_review_priority=Case(
                When(
                    source="whatsapp",
                    status=Property.Status.DRAFT,
                    published=False,
                    then=Value(0),
                ),
                default=Value(1),
                output_field=IntegerField(),
            ),
        ).order_by("whatsapp_review_priority", "-created_at")
    def perform_create(self, serializer):
        if self.request.user.is_staff:
            prop = serializer.save()
        else:
            prop = serializer.save(
                broker=user_broker(self.request.user), status=Property.Status.DRAFT,
                published=False, hidden=False, featured=False, launch=False,
                reviewed_at=None, archived_at=None, source="broker",
            )
        AuditEvent.objects.create(actor=self.request.user, action="property.created", entity_type="Property", entity_id=str(prop.id))
    def perform_update(self, serializer):
        if self.request.user.is_staff:
            prop = serializer.save()
        else:
            current = serializer.instance
            if current.status != Property.Status.DRAFT or current.published:
                raise ValidationError({"detail": "O corretor pode editar somente os próprios imóveis em rascunho."})
            prop = serializer.save(
                broker=user_broker(self.request.user), status=Property.Status.DRAFT,
                published=False, hidden=False, featured=False, launch=False,
                reviewed_at=None, archived_at=None,
            )
        AuditEvent.objects.create(actor=self.request.user, action="property.updated", entity_type="Property", entity_id=str(prop.id))

    def _ensure_editable_draft(self, prop):
        if not self.request.user.is_staff and (prop.status != Property.Status.DRAFT or prop.published):
            raise ValidationError({"detail": "O corretor pode alterar mídias somente em imóveis em rascunho."})
    def destroy(self, request, *args, **kwargs):
        prop = self.get_object()
        property_id = str(prop.id)
        stored_files = [
            (media.file.storage, media.file.name)
            for media in prop.media.all()
            if media.file and media.file.name
        ]
        with transaction.atomic():
            AuditEvent.objects.create(
                actor=request.user,
                action="property.deleted",
                entity_type="Property",
                entity_id=property_id,
                metadata={"media_count": len(stored_files)},
            )
            prop.delete()
            for file_storage, file_name in stored_files:
                transaction.on_commit(
                    lambda storage=file_storage, name=file_name: storage.delete(name)
                )
        return Response(status=status.HTTP_204_NO_CONTENT)
    @action(detail=False, methods=["post"], url_path="txt-preview")
    def txt_preview(self, request):
        upload = request.FILES.get("file")
        if not upload: raise ValidationError({"file": "Selecione um arquivo TXT."})
        if not upload.name.lower().endswith(".txt"): raise ValidationError({"file": "Envie um arquivo TXT."})
        if upload.size > 2 * 1024 * 1024: raise ValidationError({"file": "O TXT excede 2 MB."})
        raw = upload.read()
        try: text = raw.decode("utf-8")
        except UnicodeDecodeError: text = raw.decode("cp1252", errors="replace")
        suggestions = extract_property_description(text)
        values = {field: suggestion["value"] for field, suggestion in suggestions.items()}
        return Response({"values": values, "suggestions": suggestions})
    @action(detail=True, methods=["post"], url_path="media")
    def upload_media(self, request, pk=None):
        prop = self.get_object()
        self._ensure_editable_draft(prop)
        upload = request.FILES.get("file")
        if not upload: raise ValidationError({"file": "Selecione um arquivo."})
        original_extension = os.path.splitext(upload.name)[1].lower()
        if original_extension in {".heic", ".heif"}:
            try:
                upload = normalize_uploaded_image(upload, max_bytes=300 * 1024 * 1024)
            except ValidationError as exc:
                raise ValidationError({"file": exc.detail}) from exc
        extension = os.path.splitext(upload.name)[1].lower()
        kinds = {".jpg": Media.Kind.IMAGE, ".jpeg": Media.Kind.IMAGE, ".png": Media.Kind.IMAGE, ".webp": Media.Kind.IMAGE, ".mp4": Media.Kind.VIDEO, ".pdf": Media.Kind.DOCUMENT}
        if extension not in kinds: raise ValidationError({"file": "Formato não permitido."})
        if upload.size > 300 * 1024 * 1024: raise ValidationError({"file": "Arquivo excede 300 MB."})
        allowed_mime_types = {
            ".jpg": {"image/jpeg"}, ".jpeg": {"image/jpeg"},
            ".png": {"image/png"}, ".webp": {"image/webp"},
            ".mp4": {"video/mp4"}, ".pdf": {"application/pdf"},
        }
        mime_type = upload.content_type or mimetypes.guess_type(upload.name)[0] or ""
        if mime_type.lower() not in allowed_mime_types[extension]:
            raise ValidationError({"file": "O tipo do arquivo não corresponde à extensão."})
        signature = upload.read(16)
        upload.seek(0)
        valid_signature = {
            ".jpg": signature.startswith(b"\xff\xd8\xff"),
            ".jpeg": signature.startswith(b"\xff\xd8\xff"),
            ".png": signature.startswith(b"\x89PNG\r\n\x1a\n"),
            ".webp": signature.startswith(b"RIFF") and signature[8:12] == b"WEBP",
            ".mp4": signature[4:8] == b"ftyp",
            ".pdf": signature.startswith(b"%PDF-"),
        }[extension]
        if not valid_signature:
            raise ValidationError({"file": "A assinatura do arquivo é inválida ou está corrompida."})
        digest = hashlib.sha256()
        for chunk in upload.chunks(): digest.update(chunk)
        upload.seek(0)
        if prop.media.filter(sha256=digest.hexdigest()).exists(): raise ValidationError({"file": "Arquivo duplicado."})
        kind = kinds[extension]
        primary = kind == Media.Kind.IMAGE and not prop.media.filter(kind=Media.Kind.IMAGE, is_primary=True).exists()
        media = Media.objects.create(property=prop, kind=kind, file=upload, is_primary=primary, sha256=digest.hexdigest(), mime_type=mime_type, size=upload.size, status=Media.Status.READY, position=prop.media.count())
        AuditEvent.objects.create(actor=request.user, action="property.media_uploaded", entity_type="Property", entity_id=str(prop.id), metadata={"kind": kind})
        from .serializers import MediaSerializer
        return Response(MediaSerializer(media, context={"request": request}).data, status=status.HTTP_201_CREATED)
    @action(detail=True, methods=["delete"], url_path=r"media/(?P<media_id>[^/.]+)")
    def delete_media(self, request, pk=None, media_id=None):
        prop = self.get_object()
        self._ensure_editable_draft(prop)
        media = get_object_or_404(prop.media.all(), pk=media_id)
        file_name = media.file.name if media.file else ""
        file_storage = media.file.storage if media.file else None
        was_primary = media.kind == Media.Kind.IMAGE and media.is_primary
        deleted_kind = media.kind
        deleted_id = str(media.id)

        with transaction.atomic():
            media.delete()
            remaining_media = list(prop.media.order_by("position", "created_at"))
            for position, item in enumerate(remaining_media):
                if item.position != position:
                    prop.media.filter(pk=item.pk).update(position=position)
            if was_primary:
                next_image = next(
                    (
                        item for item in remaining_media
                        if item.kind == Media.Kind.IMAGE and item.status == Media.Status.READY
                    ),
                    None,
                )
                if next_image:
                    prop.media.filter(pk=next_image.pk).update(is_primary=True)
            AuditEvent.objects.create(
                actor=request.user,
                action="property.media_deleted",
                entity_type="Property",
                entity_id=str(prop.id),
                metadata={"media_id": deleted_id, "kind": deleted_kind, "was_primary": was_primary},
            )
            if file_name and file_storage:
                transaction.on_commit(lambda: file_storage.delete(file_name))

        prop._prefetched_objects_cache = {}
        return Response(self.get_serializer(prop).data)
    @action(detail=True, methods=["post"], url_path="media/(?P<media_id>[^/.]+)/primary")
    def set_primary_media(self, request, pk=None, media_id=None):
        prop = self.get_object()
        self._ensure_editable_draft(prop)
        media = prop.media.get(pk=media_id, kind=Media.Kind.IMAGE)
        with transaction.atomic():
            prop.media.filter(kind=Media.Kind.IMAGE, is_primary=True).update(is_primary=False)
            media.is_primary = True; media.save(update_fields=["is_primary", "updated_at"])
        prop._prefetched_objects_cache = {}
        return Response(self.get_serializer(prop).data)
    @action(detail=True, methods=["post"], url_path="media-order")
    def media_order(self, request, pk=None):
        prop = self.get_object()
        self._ensure_editable_draft(prop)
        ordered_ids = request.data.get("media_ids")
        if not isinstance(ordered_ids, list):
            raise ValidationError({"media_ids": "Informe a ordem das mídias."})
        current_ids = {str(media_id) for media_id in prop.media.values_list("id", flat=True)}
        if set(map(str, ordered_ids)) != current_ids:
            raise ValidationError({"media_ids": "A lista deve conter todas as mídias do imóvel."})
        with transaction.atomic():
            for position, media_id in enumerate(ordered_ids):
                prop.media.filter(pk=media_id).update(position=position)
        prop._prefetched_objects_cache = {}
        return Response(self.get_serializer(prop).data)
    @action(detail=True, methods=["post"], url_path="validate-media")
    def validate_media(self, request, pk=None):
        prop = self.get_object()
        validated = 0
        rejected = 0
        for media in prop.media.all():
            valid = bool(media.file or media.external_url) and media.kind in {
                Media.Kind.IMAGE, Media.Kind.VIDEO, Media.Kind.DOCUMENT
            }
            media.status = Media.Status.READY if valid else Media.Status.REJECTED
            media.save(update_fields=["status", "updated_at"])
            validated += int(valid)
            rejected += int(not valid)
        AuditEvent.objects.create(
            actor=request.user,
            action="property.media_validated",
            entity_type="Property",
            entity_id=str(prop.id),
            metadata={"validated": validated, "rejected": rejected},
        )
        return Response({
            "detail": f"{validated} mídia(s) validada(s).",
            "validated": validated,
            "rejected": rejected,
            "property": self.get_serializer(prop).data,
        })
    @action(detail=True, methods=["post"])
    def publish(self, request, pk=None):
        prop = self.get_object()
        try:
            with transaction.atomic(): prop.publish()
        except Exception as exc:
            raise ValidationError(getattr(exc, "message_dict", {"detail": str(exc)}))
        AuditEvent.objects.create(actor=request.user, action="property.published", entity_type="Property", entity_id=str(prop.id))
        return Response(self.get_serializer(prop).data)
    @action(detail=True, methods=["post"], url_path="toggle-featured")
    def toggle_featured(self, request, pk=None):
        prop = self.get_object()
        if not prop.is_public and not prop.featured:
            raise ValidationError({"featured": "Publique o imóvel antes de destacá-lo."})
        prop.featured = not prop.featured
        prop.save(update_fields=["featured", "updated_at"])
        AuditEvent.objects.create(actor=request.user, action="property.featured" if prop.featured else "property.unfeatured", entity_type="Property", entity_id=str(prop.id))
        return Response(self.get_serializer(prop).data)
    @action(detail=True, methods=["post"], url_path="toggle-launch")
    def toggle_launch(self, request, pk=None):
        prop = self.get_object()
        prop.launch = not prop.launch
        prop.save(update_fields=["launch", "updated_at"])
        AuditEvent.objects.create(
            actor=request.user,
            action="property.launch_featured" if prop.launch else "property.launch_unfeatured",
            entity_type="Property",
            entity_id=str(prop.id),
        )
        return Response(self.get_serializer(prop).data)
    @action(detail=True, methods=["post"], url_path="mark-in-service")
    def mark_in_service(self, request, pk=None):
        prop = self.get_object()
        prop.status = Property.Status.NEGOTIATING
        prop.save(update_fields=["status", "updated_at"])
        AuditEvent.objects.create(actor=request.user, action="property.in_service", entity_type="Property", entity_id=str(prop.id))
        return Response(self.get_serializer(prop).data)
    @action(detail=True, methods=["post"], url_path="remove-in-service")
    def remove_in_service(self, request, pk=None):
        prop = self.get_object()
        if prop.status != Property.Status.NEGOTIATING:
            raise ValidationError({"status": "Este imóvel não está em atendimento."})
        prop.status = Property.Status.AVAILABLE
        prop.save(update_fields=["status", "updated_at"])
        AuditEvent.objects.create(actor=request.user, action="property.in_service_removed", entity_type="Property", entity_id=str(prop.id))
        return Response(self.get_serializer(prop).data)
    @action(detail=True, methods=["post"], url_path="mark-sold")
    def mark_sold(self, request, pk=None):
        prop = self.get_object()
        prop.status = Property.Status.SOLD
        prop.published = False
        prop.featured = False
        prop.save(update_fields=["status", "published", "featured", "updated_at"])
        AuditEvent.objects.create(actor=request.user, action="property.sold", entity_type="Property", entity_id=str(prop.id))
        return Response(self.get_serializer(prop).data)
    @action(detail=True, methods=["post"], url_path="restore-sale")
    def restore_sale(self, request, pk=None):
        prop = self.get_object()
        if prop.status != Property.Status.SOLD:
            raise ValidationError({"status": "Este imóvel não está marcado como vendido."})
        prop.status = Property.Status.AVAILABLE
        prop.published = True
        prop.featured = False
        prop.save(update_fields=["status", "published", "featured", "updated_at"])
        AuditEvent.objects.create(actor=request.user, action="property.sale_restored", entity_type="Property", entity_id=str(prop.id))
        return Response(self.get_serializer(prop).data)
    @action(detail=True, methods=["post"], url_path="confirm-review")
    def confirm_review(self, request, pk=None):
        prop = self.get_object(); prop.reviewed_at = timezone.now(); prop.save(update_fields=["reviewed_at", "updated_at"])
        return Response(self.get_serializer(prop).data)
    @action(detail=True, methods=["post"])
    def archive(self, request, pk=None):
        prop = self.get_object()
        prop.archive()
        AuditEvent.objects.create(actor=request.user, action="property.archived", entity_type="Property", entity_id=str(prop.id))
        return Response(self.get_serializer(prop).data)
    @action(detail=True, methods=["post"], url_path="restore-archive")
    def restore_archive(self, request, pk=None):
        prop = self.get_object()
        if prop.status != Property.Status.ARCHIVED or not prop.archived_at:
            raise ValidationError({"status": "Este imóvel não está arquivado."})
        prop.restore_archive()
        AuditEvent.objects.create(actor=request.user, action="property.archive_restored", entity_type="Property", entity_id=str(prop.id))
        return Response(self.get_serializer(prop).data)

class AdminDevelopmentViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAdminUser]; serializer_class = DevelopmentSerializer; queryset = Development.objects.all()
class AdminLeadViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAdminUser]; serializer_class = AdminLeadSerializer; queryset = Lead.objects.all().select_related("property", "development", "broker")


def crm_contacts_owned_by(broker):
    return CRMContact.objects.filter(
        Q(assigned_broker=broker)
        | Q(
            opportunities__broker=broker,
            opportunities__stage__in=[
                CRMOpportunity.Stage.NEW,
                CRMOpportunity.Stage.CONTACTED,
                CRMOpportunity.Stage.VISIT,
                CRMOpportunity.Stage.PROPOSAL,
                CRMOpportunity.Stage.NEGOTIATION,
                CRMOpportunity.Stage.PAUSED,
            ],
        )
        | Q(tasks__broker=broker, tasks__status=CRMTask.Status.PENDING)
    ).distinct()


def crm_contacts_for(user):
    if can_view_all_crm(user):
        return CRMContact.objects.all()
    return crm_contacts_owned_by(user_broker(user))


def _edit_distance(first, second):
    previous = list(range(len(second) + 1))
    for first_index, first_character in enumerate(first, start=1):
        diagonal = previous[0]
        previous[0] = first_index
        for second_index, second_character in enumerate(second, start=1):
            above = previous[second_index]
            previous[second_index] = min(
                previous[second_index] + 1,
                previous[second_index - 1] + 1,
                diagonal + (first_character != second_character),
            )
            diagonal = above
    return previous[-1]


def _fuzzy_name_match(search_name, query):
    if query in search_name:
        return True
    name_tokens = search_name.split()
    for query_token in query.split():
        tolerance = 2 if len(query_token) >= 7 else 1 if len(query_token) >= 4 else 0
        if not any(_edit_distance(query_token, name_token) <= tolerance for name_token in name_tokens):
            return False
    return True


def filter_contacts_by_search(queryset, raw_search):
    raw_search = (raw_search or "").strip()[:80]
    if not raw_search:
        return queryset
    normalized_search = person_name_search_key(raw_search)
    direct_filter = (
        Q(name__icontains=raw_search)
        | Q(document__icontains=raw_search)
        | Q(phone__icontains=raw_search)
        | Q(email__icontains=raw_search)
        | Q(city__icontains=raw_search)
    )
    fuzzy_ids = []
    if normalized_search:
        direct_filter |= Q(search_name__contains=normalized_search)
        fuzzy_ids = [
            contact_id
            for contact_id, search_name in queryset.order_by().values_list("id", "search_name")
            if _fuzzy_name_match(search_name, normalized_search)
        ]
    return queryset.filter(direct_filter | Q(id__in=fuzzy_ids))


def contact_holder_brokers(contact):
    return Broker.objects.filter(
        Q(crm_contacts=contact)
        | Q(
            crm_opportunities__contact=contact,
            crm_opportunities__stage__in=[
                CRMOpportunity.Stage.NEW,
                CRMOpportunity.Stage.CONTACTED,
                CRMOpportunity.Stage.VISIT,
                CRMOpportunity.Stage.PROPOSAL,
                CRMOpportunity.Stage.NEGOTIATION,
                CRMOpportunity.Stage.PAUSED,
            ],
        )
        | Q(
            crm_tasks__contact=contact,
            crm_tasks__status=CRMTask.Status.PENDING,
        )
    ).distinct().order_by("name", "id")


def crm_opportunities_for(user):
    queryset = CRMOpportunity.objects.all()
    return queryset if can_view_all_crm(user) else queryset.filter(broker=user_broker(user))


def crm_tasks_for(user):
    queryset = CRMTask.objects.all()
    return queryset if can_view_all_crm(user) else queryset.filter(broker=user_broker(user))


def notify_broker(broker, *, kind, title, message="", link="/admin", priority=CRMNotification.Priority.NORMAL, unique_key=None, source_task=None):
    if not broker or not broker.active or not broker.user_id or not broker.user.is_active:
        return None
    notification, _ = CRMNotification.objects.get_or_create(
        unique_key=unique_key,
        defaults={
            "recipient": broker.user,
            "broker": broker,
            "source_task": source_task,
            "kind": kind,
            "priority": priority,
            "title": title,
            "message": message,
            "link": link,
        },
    ) if unique_key else (CRMNotification.objects.create(
        recipient=broker.user, broker=broker, source_task=source_task, kind=kind,
        priority=priority, title=title, message=message, link=link,
    ), True)
    return notification


def ensure_due_notifications(user):
    broker = user_broker(user)
    if not broker:
        return
    now = timezone.now()
    limit = now + timedelta(hours=24)
    tasks = CRMTask.objects.filter(broker=broker, status=CRMTask.Status.PENDING, due_at__lte=limit).select_related("contact")
    for task in tasks:
        overdue = task.due_at < now
        if overdue:
            CRMNotification.objects.filter(
                source_task=task,
                kind=CRMNotification.Kind.TASK_DUE,
                read_at__isnull=True,
            ).update(read_at=now)
        notify_broker(
            broker,
            kind=CRMNotification.Kind.TASK_OVERDUE if overdue else CRMNotification.Kind.TASK_DUE,
            title="Tarefa atrasada" if overdue else "Tarefa nas próximas 24 horas",
            message=f"{task.title} · {task.contact.name}",
            priority=CRMNotification.Priority.HIGH if overdue else CRMNotification.Priority.NORMAL,
            unique_key=f"task:{task.id}:{'overdue' if overdue else 'due'}",
            source_task=task,
        )


class CRMAuditMixin:
    audit_entity = "CRM"
    broker_assignment_field = None

    def _restricted_broker(self):
        return None if can_view_all_crm(self.request.user) else user_broker(self.request.user)

    def perform_create(self, serializer):
        defaults = {}
        broker = self._restricted_broker()
        if broker and self.broker_assignment_field:
            defaults[self.broker_assignment_field] = broker
        instance = serializer.save(**defaults)
        AuditEvent.objects.create(
            actor=self.request.user,
            action=f"crm.{self.audit_entity.casefold()}.created",
            entity_type=self.audit_entity,
            entity_id=str(instance.id),
        )

    def perform_update(self, serializer):
        broker = self._restricted_broker()
        if broker and self.broker_assignment_field in serializer.validated_data:
            assigned = serializer.validated_data[self.broker_assignment_field]
            if assigned != broker:
                raise ValidationError({self.broker_assignment_field: "O corretor não pode transferir registros para outro usuário."})
        instance = serializer.save()
        AuditEvent.objects.create(
            actor=self.request.user,
            action=f"crm.{self.audit_entity.casefold()}.updated",
            entity_type=self.audit_entity,
            entity_id=str(instance.id),
        )


class AdminCRMContactViewSet(CRMAuditMixin, viewsets.ModelViewSet):
    permission_classes = [IsCRMUser]
    serializer_class = CRMContactSerializer
    pagination_class = CRMContactPagination
    filter_backends = (DjangoFilterBackend, OrderingFilter)
    audit_entity = "CRMContact"
    broker_assignment_field = "assigned_broker"
    filterset_fields = ("status", "profile", "person_type", "assigned_broker", "source")
    search_fields = ("name", "document", "phone", "email", "city", "tags")
    ordering_fields = ("name", "created_at", "updated_at", "last_contact_at")

    def get_queryset(self):
        queryset = crm_contacts_for(self.request.user).select_related("assigned_broker").prefetch_related(
            "property_links__property"
        ).annotate(
            opportunity_count=Count("opportunities", distinct=True),
            pending_task_count=Count("tasks", filter=Q(tasks__status=CRMTask.Status.PENDING), distinct=True),
        ).order_by("name", "created_at")
        return filter_contacts_by_search(queryset, self.request.query_params.get("search"))

    def perform_update(self, serializer):
        old_broker = serializer.instance.assigned_broker
        broker_changed = "assigned_broker" in serializer.validated_data
        new_broker = serializer.validated_data.get("assigned_broker", old_broker)
        with transaction.atomic():
            super().perform_update(serializer)
            if broker_changed and old_broker != new_broker:
                active_opportunities = serializer.instance.opportunities.exclude(
                    stage__in=[
                        CRMOpportunity.Stage.WON,
                        CRMOpportunity.Stage.LOST,
                        CRMOpportunity.Stage.RELEASED,
                    ],
                )
                pending_tasks = serializer.instance.tasks.filter(status=CRMTask.Status.PENDING)
                if old_broker:
                    active_opportunities = active_opportunities.filter(broker=old_broker)
                    pending_tasks = pending_tasks.filter(broker=old_broker)
                else:
                    active_opportunities = active_opportunities.filter(broker__isnull=True)
                    pending_tasks = pending_tasks.filter(broker__isnull=True)
                if new_broker:
                    assigned_opportunity_count = active_opportunities.update(broker=new_broker)
                    pending_tasks.update(broker=new_broker)
                    if not old_broker and not assigned_opportunity_count:
                        released_opportunity = serializer.instance.opportunities.filter(
                            broker__isnull=True,
                            stage=CRMOpportunity.Stage.RELEASED,
                        ).order_by("-updated_at").first()
                        if released_opportunity:
                            released_opportunity.broker = new_broker
                            released_opportunity.stage = CRMOpportunity.Stage.NEW
                            released_opportunity.save(update_fields=["broker", "stage", "updated_at"])
                else:
                    active_opportunities.update(
                        broker=None,
                        stage=CRMOpportunity.Stage.RELEASED,
                    )
                    pending_tasks.update(broker=None, status=CRMTask.Status.CANCELED)

    @action(detail=False, methods=["get"], url_path="summary")
    def summary(self, request):
        return Response({
            "contacts": crm_contacts_for(request.user).count(),
            "open_opportunities": crm_opportunities_for(request.user).exclude(
                stage__in=[
                    CRMOpportunity.Stage.WON,
                    CRMOpportunity.Stage.LOST,
                    CRMOpportunity.Stage.RELEASED,
                ],
            ).count(),
            "pending_follow_ups": crm_tasks_for(request.user).filter(
                status=CRMTask.Status.PENDING,
            ).count(),
            "won_opportunities": crm_opportunities_for(request.user).filter(
                stage=CRMOpportunity.Stage.WON,
            ).count(),
        })

    @action(detail=False, methods=["get"], url_path="choices")
    def choices(self, request):
        queryset = filter_contacts_by_search(
            crm_contacts_for(request.user).order_by("name", "created_at"),
            request.query_params.get("search"),
        )
        page = self.paginate_queryset(queryset)
        serializer = CRMContactChoiceSerializer(page if page is not None else queryset, many=True)
        return self.get_paginated_response(serializer.data) if page is not None else Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="available")
    def available(self, request):
        queryset = CRMContact.objects.filter(
            assigned_broker__isnull=True,
            status=CRMContact.Status.ACTIVE,
        ).exclude(
            opportunities__in=CRMOpportunity.objects.exclude(
                broker__isnull=True,
            ).exclude(stage__in=[
                CRMOpportunity.Stage.WON,
                CRMOpportunity.Stage.LOST,
                CRMOpportunity.Stage.RELEASED,
            ]),
        ).exclude(
            tasks__in=CRMTask.objects.filter(
                broker__isnull=False,
                status=CRMTask.Status.PENDING,
            ),
        ).distinct().order_by("name", "created_at")
        queryset = filter_contacts_by_search(queryset, request.query_params.get("search"))
        page = self.paginate_queryset(queryset)
        serializer = CRMContactPoolSerializer(page if page is not None else queryset, many=True)
        return self.get_paginated_response(serializer.data) if page is not None else Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="claim")
    def claim(self, request, pk=None):
        broker = user_broker(request.user)
        if not broker:
            raise ValidationError({"detail": "Somente um corretor ou gestor comercial pode assumir atendimentos."})
        with transaction.atomic():
            contact = get_object_or_404(
                CRMContact.objects.select_for_update(),
                pk=pk,
                status=CRMContact.Status.ACTIVE,
            )
            if contact.assigned_broker_id and contact.assigned_broker_id != broker.id:
                raise ValidationError({"detail": "Este contato já foi assumido por outro corretor."})
            identity_filter = Q()
            if contact.document:
                identity_filter |= Q(document=contact.document)
            if contact.normalized_phone:
                identity_filter |= Q(normalized_phone=contact.normalized_phone)
            if contact.normalized_email:
                identity_filter |= Q(normalized_email=contact.normalized_email)
            if identity_filter and crm_contacts_owned_by(broker).exclude(pk=contact.pk).filter(identity_filter).exists():
                raise ValidationError({"detail": "Este cliente já está cadastrado na sua carteira."})
            already_assigned = contact.assigned_broker_id == broker.id
            if not already_assigned:
                contact.assigned_broker = broker
                contact.save(update_fields=["assigned_broker", "updated_at"])
            opportunity = contact.opportunities.select_for_update().filter(
                broker__isnull=True,
            ).exclude(
                stage__in=[CRMOpportunity.Stage.WON, CRMOpportunity.Stage.LOST],
            ).order_by("created_at").first()
            if opportunity:
                opportunity.broker = broker
                update_fields = ["broker", "updated_at"]
                if opportunity.stage == CRMOpportunity.Stage.RELEASED:
                    opportunity.stage = CRMOpportunity.Stage.NEW
                    update_fields.append("stage")
                opportunity.save(update_fields=update_fields)
            elif not contact.opportunities.filter(
                broker=broker,
            ).exclude(stage__in=[
                CRMOpportunity.Stage.WON,
                CRMOpportunity.Stage.LOST,
                CRMOpportunity.Stage.RELEASED,
            ]).exists():
                opportunity = CRMOpportunity.objects.create(
                    contact=contact,
                    broker=broker,
                    title=f"Atendimento de {contact.name}"[:200],
                    stage=CRMOpportunity.Stage.NEW,
                    source="shared_pool",
                )
            if not already_assigned:
                CRMActivity.objects.create(
                    contact=contact,
                    opportunity=opportunity,
                    actor=request.user,
                    kind=CRMActivity.Kind.NOTE,
                    description=f"Atendimento assumido por {broker.name}.",
                    metadata={"event": "contact_claimed", "broker_id": str(broker.id)},
                )
                AuditEvent.objects.create(
                    actor=request.user,
                    action="crm.contact.claimed",
                    entity_type="CRMContact",
                    entity_id=str(contact.id),
                    metadata={"broker_id": str(broker.id)},
                )
        return Response({"detail": "Contato adicionado à sua carteira.", "contact_id": str(contact.id)})

    @action(detail=True, methods=["post"], url_path="release")
    def release(self, request, pk=None):
        if not request.user.is_staff:
            raise PermissionDenied("Somente administradores podem remover contatos da carteira de um corretor.")
        with transaction.atomic():
            contact = get_object_or_404(
                CRMContact.objects.select_for_update(),
                pk=pk,
            )
            holders = contact_holder_brokers(contact)
            requested_broker = request.data.get("broker")
            if requested_broker:
                broker = get_object_or_404(holders, pk=requested_broker)
            else:
                holder_ids = list(holders.values_list("id", flat=True)[:2])
                if not holder_ids:
                    raise ValidationError({"detail": "Este contato não está atribuído a nenhum corretor."})
                if len(holder_ids) > 1:
                    raise ValidationError({"detail": "Informe qual corretor deve ser removido deste contato."})
                broker = holders.get(pk=holder_ids[0])
            open_opportunities = contact.opportunities.filter(broker=broker).exclude(
                stage__in=[
                    CRMOpportunity.Stage.WON,
                    CRMOpportunity.Stage.LOST,
                    CRMOpportunity.Stage.RELEASED,
                ],
            )
            pending_tasks = contact.tasks.filter(broker=broker, status=CRMTask.Status.PENDING)
            opportunity_count = open_opportunities.update(
                broker=None,
                stage=CRMOpportunity.Stage.RELEASED,
            )
            task_count = pending_tasks.update(broker=None, status=CRMTask.Status.CANCELED)
            if contact.assigned_broker_id == broker.id:
                contact.assigned_broker = None
                contact.save(update_fields=["assigned_broker", "updated_at"])
            returned_to_pool = not contact_holder_brokers(contact).exists()
            CRMActivity.objects.create(
                contact=contact,
                actor=request.user,
                kind=CRMActivity.Kind.NOTE,
                description=f"Contato removido da carteira de {broker.name} pelo administrador.",
                metadata={
                    "event": "contact_released",
                    "broker_id": str(broker.id),
                    "open_opportunities_released": opportunity_count,
                    "pending_tasks_released": task_count,
                    "returned_to_pool": returned_to_pool,
                },
            )
            AuditEvent.objects.create(
                actor=request.user,
                action="crm.contact.released",
                entity_type="CRMContact",
                entity_id=str(contact.id),
                metadata={
                    "broker_id": str(broker.id),
                    "open_opportunities_released": opportunity_count,
                    "pending_tasks_released": task_count,
                    "returned_to_pool": returned_to_pool,
                },
            )
        return Response({
            "detail": (
                "Contato removido da carteira e devolvido aos leads disponíveis."
                if returned_to_pool
                else "Contato removido deste corretor; ainda existe outro atendimento aberto."
            ),
            "contact_id": str(contact.id),
            "returned_to_pool": returned_to_pool,
        })

    @action(detail=True, methods=["get"], url_path="holders")
    def holders(self, request, pk=None):
        if not request.user.is_staff:
            raise PermissionDenied("Somente administradores podem consultar atribuições do contato.")
        contact = self.get_object()
        return Response([
            {
                "id": str(broker.id),
                "name": broker.name,
                "username": broker.user.username if broker.user_id else "",
            }
            for broker in contact_holder_brokers(contact).select_related("user")
        ])


class AdminCRMPropertyLinkViewSet(CRMAuditMixin, viewsets.ModelViewSet):
    permission_classes = [IsCRMUser]
    serializer_class = CRMPropertyLinkSerializer
    audit_entity = "CRMPropertyLink"
    filterset_fields = ("contact", "property", "relationship", "active")
    search_fields = ("contact__name", "property__title", "development_name", "unit_reference")

    def get_queryset(self):
        return CRMPropertyLink.objects.filter(contact__in=crm_contacts_for(self.request.user)).select_related("contact", "property")

    def perform_create(self, serializer):
        if not crm_contacts_for(self.request.user).filter(pk=serializer.validated_data["contact"].pk).exists():
            raise ValidationError({"contact": "Contato fora da sua carteira."})
        super().perform_create(serializer)

    def perform_update(self, serializer):
        contact = serializer.validated_data.get("contact", serializer.instance.contact)
        if not crm_contacts_for(self.request.user).filter(pk=contact.pk).exists():
            raise ValidationError({"contact": "Contato fora da sua carteira."})
        super().perform_update(serializer)


class AdminCRMOpportunityViewSet(CRMAuditMixin, viewsets.ModelViewSet):
    permission_classes = [IsCRMUser]
    serializer_class = CRMOpportunitySerializer
    audit_entity = "CRMOpportunity"
    broker_assignment_field = "broker"
    filterset_fields = ("stage", "broker", "contact", "property", "source")
    search_fields = ("title", "contact__name", "property__title", "notes")
    ordering_fields = ("created_at", "updated_at", "next_action_at", "expected_value")

    def get_queryset(self):
        return crm_opportunities_for(self.request.user).select_related("contact", "property", "broker").annotate(
            proposal_count=Count("proposals", distinct=True)
        ).order_by("-updated_at")

    def perform_create(self, serializer):
        contact = serializer.validated_data["contact"]
        if not crm_contacts_for(self.request.user).filter(pk=contact.pk).exists():
            raise ValidationError({"contact": "Contato fora da sua carteira."})
        super().perform_create(serializer)
        opportunity = serializer.instance
        notify_broker(
            opportunity.broker,
            kind=CRMNotification.Kind.ASSIGNMENT,
            title="Nova oportunidade atribuída",
            message=f"{opportunity.contact.name} · {opportunity.title}",
            unique_key=f"opportunity:{opportunity.id}:assigned:{opportunity.broker_id}",
        )

    def perform_update(self, serializer):
        old_broker_id = serializer.instance.broker_id
        contact = serializer.validated_data.get("contact", serializer.instance.contact)
        if not crm_contacts_for(self.request.user).filter(pk=contact.pk).exists():
            raise ValidationError({"contact": "Contato fora da sua carteira."})
        super().perform_update(serializer)
        opportunity = serializer.instance
        if opportunity.broker_id and opportunity.broker_id != old_broker_id:
            notify_broker(
                opportunity.broker,
                kind=CRMNotification.Kind.ASSIGNMENT,
                title="Oportunidade transferida para você",
                message=f"{opportunity.contact.name} · {opportunity.title}",
                unique_key=f"opportunity:{opportunity.id}:assigned:{opportunity.broker_id}",
            )


class AdminCRMTaskViewSet(CRMAuditMixin, viewsets.ModelViewSet):
    permission_classes = [IsCRMUser]
    serializer_class = CRMTaskSerializer
    audit_entity = "CRMTask"
    broker_assignment_field = "broker"
    filterset_fields = ("status", "kind", "broker", "contact", "opportunity")
    search_fields = ("title", "description", "contact__name")
    ordering_fields = ("due_at", "created_at")

    def get_queryset(self):
        return crm_tasks_for(self.request.user).select_related("contact", "opportunity", "property", "broker")

    def perform_create(self, serializer):
        contact = serializer.validated_data["contact"]
        opportunity = serializer.validated_data.get("opportunity")
        if not crm_contacts_for(self.request.user).filter(pk=contact.pk).exists():
            raise ValidationError({"contact": "Contato fora da sua carteira."})
        if opportunity and not crm_opportunities_for(self.request.user).filter(pk=opportunity.pk).exists():
            raise ValidationError({"opportunity": "Oportunidade fora da sua carteira."})
        super().perform_create(serializer)
        task = serializer.instance
        notify_broker(
            task.broker,
            kind=CRMNotification.Kind.ASSIGNMENT,
            title="Nova tarefa atribuída",
            message=f"{task.title} · {task.contact.name}",
            unique_key=f"task:{task.id}:assigned",
            source_task=task,
        )

    def perform_update(self, serializer):
        contact = serializer.validated_data.get("contact", serializer.instance.contact)
        opportunity = serializer.validated_data.get("opportunity", serializer.instance.opportunity)
        if not crm_contacts_for(self.request.user).filter(pk=contact.pk).exists():
            raise ValidationError({"contact": "Contato fora da sua carteira."})
        if opportunity and not crm_opportunities_for(self.request.user).filter(pk=opportunity.pk).exists():
            raise ValidationError({"opportunity": "Oportunidade fora da sua carteira."})
        due_changed = "due_at" in serializer.validated_data
        super().perform_update(serializer)
        task = serializer.instance
        task_notifications = CRMNotification.objects.filter(
            source_task=task,
            kind__in=[CRMNotification.Kind.TASK_DUE, CRMNotification.Kind.TASK_OVERDUE],
        )
        if task.status != CRMTask.Status.PENDING:
            task_notifications.filter(read_at__isnull=True).update(read_at=timezone.now())
        elif due_changed:
            task_notifications.delete()


class AdminCRMActivityViewSet(CRMAuditMixin, viewsets.ModelViewSet):
    permission_classes = [IsCRMUser]
    serializer_class = CRMActivitySerializer
    audit_entity = "CRMActivity"
    filterset_fields = ("contact", "opportunity", "kind")
    search_fields = ("description", "contact__name")

    def get_queryset(self):
        return CRMActivity.objects.filter(contact__in=crm_contacts_for(self.request.user)).select_related("contact", "opportunity", "actor")

    def perform_create(self, serializer):
        if not crm_contacts_for(self.request.user).filter(pk=serializer.validated_data["contact"].pk).exists():
            raise ValidationError({"contact": "Contato fora da sua carteira."})
        instance = serializer.save(actor=self.request.user)
        AuditEvent.objects.create(
            actor=self.request.user,
            action="crm.activity.created",
            entity_type="CRMActivity",
            entity_id=str(instance.id),
        )

    def perform_update(self, serializer):
        contact = serializer.validated_data.get("contact", serializer.instance.contact)
        if not crm_contacts_for(self.request.user).filter(pk=contact.pk).exists():
            raise ValidationError({"contact": "Contato fora da sua carteira."})
        super().perform_update(serializer)


class AdminCRMProposalViewSet(CRMAuditMixin, viewsets.ModelViewSet):
    permission_classes = [IsCRMUser]
    serializer_class = CRMProposalSerializer
    audit_entity = "CRMProposal"
    filterset_fields = ("status", "opportunity")
    ordering_fields = ("created_at", "total_value", "valid_until")

    def get_queryset(self):
        return CRMProposal.objects.filter(opportunity__in=crm_opportunities_for(self.request.user)).select_related("opportunity__contact", "opportunity__property")

    def perform_create(self, serializer):
        opportunity = serializer.validated_data["opportunity"]
        if not crm_opportunities_for(self.request.user).filter(pk=opportunity.pk).exists():
            raise ValidationError({"opportunity": "Oportunidade fora da sua carteira."})
        super().perform_create(serializer)

    def perform_update(self, serializer):
        opportunity = serializer.validated_data.get("opportunity", serializer.instance.opportunity)
        if not crm_opportunities_for(self.request.user).filter(pk=opportunity.pk).exists():
            raise ValidationError({"opportunity": "Oportunidade fora da sua carteira."})
        super().perform_update(serializer)


class AdminCRMImportRowViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCRMManager]
    serializer_class = CRMImportRowSerializer
    queryset = CRMImportRow.objects.select_related("batch", "matched_contact", "matched_property")
    http_method_names = ("get", "patch", "head", "options")
    filterset_fields = ("batch", "status", "matched_contact", "matched_property")

    def partial_update(self, request, *args, **kwargs):
        row = self.get_object()
        if row.batch.status != CRMImportBatch.Status.REVIEW:
            raise ValidationError({"detail": "Esta importação não está mais em revisão."})
        allowed = {"normalized_data", "matched_property", "status"}
        unexpected = set(request.data) - allowed
        if unexpected:
            raise ValidationError({"detail": "Somente os dados normalizados, o imóvel e a situação podem ser revisados."})
        data = request.data.get("normalized_data", row.normalized_data)
        normalized, errors = sanitize_import_row(data)
        row.matched_contact = find_duplicate_contact(normalized)
        duplicate_row = find_duplicate_import_row(row.batch, normalized, exclude_id=row.id)
        if duplicate_row:
            normalized["duplicate_of_row"] = duplicate_row.row_number
        requested_status = request.data.get("status")
        if requested_status == CRMImportRow.Status.IGNORED:
            row.status = CRMImportRow.Status.IGNORED
        elif errors:
            row.status = CRMImportRow.Status.ERROR
        else:
            row.status = CRMImportRow.Status.DUPLICATE if row.matched_contact or duplicate_row else CRMImportRow.Status.VALID
        row.normalized_data = normalized
        row.errors = errors
        if "matched_property" in request.data:
            property_id = request.data.get("matched_property")
            row.matched_property = get_object_or_404(Property, pk=property_id) if property_id else None
        row.save(update_fields=["normalized_data", "errors", "status", "matched_contact", "matched_property", "updated_at"])
        counts = row.batch.rows.values("status").annotate(total=Count("id"))
        totals = {item["status"]: item["total"] for item in counts}
        row.batch.valid_rows = totals.get(CRMImportRow.Status.VALID, 0)
        row.batch.duplicate_rows = totals.get(CRMImportRow.Status.DUPLICATE, 0)
        row.batch.error_rows = totals.get(CRMImportRow.Status.ERROR, 0)
        row.batch.save(update_fields=["valid_rows", "duplicate_rows", "error_rows", "updated_at"])
        return Response(self.get_serializer(row).data)


class AdminCRMImportBatchViewSet(mixins.CreateModelMixin, viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsCRMManager]
    serializer_class = CRMImportBatchSerializer
    queryset = CRMImportBatch.objects.select_related("created_by")

    def create(self, request, *args, **kwargs):
        upload = request.FILES.get("file")
        if not upload:
            raise ValidationError({"file": "Selecione um arquivo CSV ou PDF."})
        try:
            validate_crm_import(upload)
        except ValueError as exc:
            raise ValidationError({"file": str(exc)}) from exc
        source_hash = file_sha256(upload)
        batch = CRMImportBatch.objects.create(
            file=upload,
            original_name=upload.name,
            source_hash=source_hash,
            source_label=request.data.get("source_label", ""),
            created_by=request.user,
        )
        process_import_batch(batch)
        AuditEvent.objects.create(
            actor=request.user,
            action="crm.import.previewed",
            entity_type="CRMImportBatch",
            entity_id=str(batch.id),
            metadata={"rows": batch.total_rows, "status": batch.status},
        )
        return Response(self.get_serializer(batch).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["post"], url_path="ignore-invalid")
    def ignore_invalid(self, request, pk=None):
        batch = self.get_object()
        if batch.status != CRMImportBatch.Status.REVIEW:
            raise ValidationError({"detail": "Esta importação não está mais em revisão."})
        ignored = batch.rows.filter(status=CRMImportRow.Status.ERROR).update(status=CRMImportRow.Status.IGNORED)
        batch.error_rows = 0
        batch.save(update_fields=["error_rows", "updated_at"])
        AuditEvent.objects.create(
            actor=request.user,
            action="crm.import.invalid_rows_ignored",
            entity_type="CRMImportBatch",
            entity_id=str(batch.id),
            metadata={"ignored_rows": ignored},
        )
        return Response({"ignored_rows": ignored, "batch": self.get_serializer(batch).data})

    @action(detail=True, methods=["post"])
    def commit(self, request, pk=None):
        batch = self.get_object()
        try:
            imported = commit_import_batch(batch, request.user)
        except ValueError as exc:
            raise ValidationError({"detail": str(exc)}) from exc
        AuditEvent.objects.create(
            actor=request.user,
            action="crm.import.committed",
            entity_type="CRMImportBatch",
            entity_id=str(batch.id),
            metadata={"imported_rows": imported},
        )
        return Response({"imported_rows": imported, "batch": self.get_serializer(batch).data})


class CRMNotificationViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsCRMUser]
    serializer_class = CRMNotificationSerializer

    def get_queryset(self):
        ensure_due_notifications(self.request.user)
        return CRMNotification.objects.filter(recipient=self.request.user).select_related("broker", "source_task")

    @action(detail=True, methods=["post"], url_path="mark-read")
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        if not notification.read_at:
            notification.read_at = timezone.now()
            notification.save(update_fields=["read_at", "updated_at"])
        return Response(self.get_serializer(notification).data)

    @action(detail=False, methods=["post"], url_path="mark-all-read")
    def mark_all_read(self, request):
        updated = CRMNotification.objects.filter(recipient=request.user, read_at__isnull=True).update(read_at=timezone.now())
        return Response({"updated": updated})


class CRMPropertyReferenceView(APIView):
    permission_classes = [IsCRMUser]

    def get(self, request):
        properties = Property.objects.filter(archived_at__isnull=True).order_by("title").values(
            "id", "public_id", "title", "status", "city", "neighborhood", "price"
        )
        return Response(list(properties))


class CRMTeamReferenceView(APIView):
    permission_classes = [IsCRMUser]

    def get(self, request):
        brokers = Broker.objects.filter(active=True).order_by("name")
        if not can_view_all_crm(request.user):
            brokers = brokers.filter(pk=user_broker(request.user).pk)
        return Response(list(brokers.values("id", "name", "role")))


class CRMReportView(APIView):
    permission_classes = [IsCRMUser]

    def get(self, request):
        today = timezone.localdate()
        start_date = parse_date(request.query_params.get("date_from", "")) or today - timedelta(days=29)
        end_date = parse_date(request.query_params.get("date_to", "")) or today
        if start_date > end_date:
            raise ValidationError({"date_to": "A data final deve ser posterior à data inicial."})
        if (end_date - start_date).days > 366:
            raise ValidationError({"date_to": "O relatório permite no máximo 367 dias por consulta."})
        start = timezone.make_aware(datetime.combine(start_date, time.min))
        end = timezone.make_aware(datetime.combine(end_date + timedelta(days=1), time.min))

        opportunities = crm_opportunities_for(request.user).select_related("broker")
        period_opportunities = opportunities.filter(created_at__gte=start, created_at__lt=end)
        closed = opportunities.filter(closed_at__gte=start, closed_at__lt=end)
        won = closed.filter(stage=CRMOpportunity.Stage.WON)
        lost = closed.filter(stage=CRMOpportunity.Stage.LOST)
        decided = won.count() + lost.count()
        open_stages = [
            CRMOpportunity.Stage.NEW, CRMOpportunity.Stage.CONTACTED, CRMOpportunity.Stage.VISIT,
            CRMOpportunity.Stage.PROPOSAL, CRMOpportunity.Stage.NEGOTIATION, CRMOpportunity.Stage.PAUSED,
        ]
        pipeline = opportunities.filter(stage__in=open_stages)
        contacts = crm_contacts_for(request.user).filter(created_at__gte=start, created_at__lt=end)
        tasks = crm_tasks_for(request.user)
        proposals = CRMProposal.objects.filter(opportunity__in=opportunities)

        cycle_days = [
            (closed_at - created_at).total_seconds() / 86400
            for created_at, closed_at in won.values_list("created_at", "closed_at") if closed_at
        ]
        trend_rows = period_opportunities.annotate(day=TruncDate("created_at")).values("day").annotate(total=Count("id")).order_by("day")
        won_trend_rows = won.annotate(day=TruncDate("closed_at")).values("day").annotate(total=Count("id")).order_by("day")
        trend = {}
        cursor = start_date
        while cursor <= end_date:
            trend[cursor.isoformat()] = {"date": cursor.isoformat(), "opportunities": 0, "won": 0}
            cursor += timedelta(days=1)
        for row in trend_rows:
            trend[row["day"].isoformat()]["opportunities"] = row["total"]
        for row in won_trend_rows:
            trend[row["day"].isoformat()]["won"] = row["total"]

        source_rows = period_opportunities.values("source").annotate(total=Count("id")).order_by("-total", "source")
        stage_rows = opportunities.exclude(
            stage=CRMOpportunity.Stage.RELEASED,
        ).values("stage").annotate(total=Count("id"), value=Sum("expected_value")).order_by("stage")
        loss_rows = lost.exclude(loss_reason="").values("loss_reason").annotate(total=Count("id")).order_by("-total")[:10]

        brokers = Broker.objects.filter(active=True)
        if not can_view_all_crm(request.user):
            brokers = brokers.filter(pk=user_broker(request.user).pk)
        broker_performance = []
        for broker in brokers.order_by("name"):
            broker_created = period_opportunities.filter(broker=broker)
            broker_closed = closed.filter(broker=broker)
            broker_won = broker_closed.filter(stage=CRMOpportunity.Stage.WON).count()
            broker_lost = broker_closed.filter(stage=CRMOpportunity.Stage.LOST).count()
            broker_decided = broker_won + broker_lost
            broker_performance.append({
                "broker_id": str(broker.id),
                "broker_name": broker.name,
                "opportunities": broker_created.count(),
                "won": broker_won,
                "lost": broker_lost,
                "conversion_rate": round(broker_won * 100 / broker_decided, 1) if broker_decided else 0,
                "won_value": float(broker_closed.filter(stage=CRMOpportunity.Stage.WON).aggregate(value=Sum("expected_value"))["value"] or 0),
                "overdue_tasks": tasks.filter(broker=broker, status=CRMTask.Status.PENDING, due_at__lt=timezone.now()).count(),
            })

        payload = {
            "period": {"date_from": start_date, "date_to": end_date},
            "metrics": {
                "new_contacts": contacts.count(),
                "new_opportunities": period_opportunities.count(),
                "won": won.count(),
                "lost": lost.count(),
                "conversion_rate": round(won.count() * 100 / decided, 1) if decided else 0,
                "pipeline_value": float(pipeline.aggregate(value=Sum("expected_value"))["value"] or 0),
                "won_value": float(won.aggregate(value=Sum("expected_value"))["value"] or 0),
                "average_cycle_days": round(sum(cycle_days) / len(cycle_days), 1) if cycle_days else 0,
                "overdue_tasks": tasks.filter(status=CRMTask.Status.PENDING, due_at__lt=timezone.now()).count(),
                "completed_visits": tasks.filter(kind=CRMTask.Kind.VISIT, status=CRMTask.Status.COMPLETED, completed_at__gte=start, completed_at__lt=end).count(),
                "sent_proposals": proposals.filter(status__in=[CRMProposal.Status.SENT, CRMProposal.Status.ANALYSIS, CRMProposal.Status.COUNTER, CRMProposal.Status.ACCEPTED], updated_at__gte=start, updated_at__lt=end).count(),
            },
            "by_stage": [{"stage": row["stage"], "label": dict(CRMOpportunity.Stage.choices)[row["stage"]], "total": row["total"], "value": float(row["value"] or 0)} for row in stage_rows],
            "by_source": list(source_rows),
            "loss_reasons": list(loss_rows),
            "broker_performance": broker_performance,
            "trend": list(trend.values()),
        }
        export_format = request.query_params.get("export", "").lower()
        if export_format == "xlsx":
            return self._xlsx_response(payload)
        if export_format == "pdf":
            return self._pdf_response(payload)
        if export_format:
            raise ValidationError({"export": "Formato inválido. Use xlsx ou pdf."})
        return Response(payload)

    @staticmethod
    def _filename(payload, extension):
        return f"relatorio-crm-{payload['period']['date_from']}-a-{payload['period']['date_to']}.{extension}"

    @classmethod
    def _xlsx_response(cls, payload):
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Resumo"
        navy = "0B2947"
        gold = "C89B3C"

        def write_sheet(sheet, headers, rows):
            sheet.append(headers)
            for cell in sheet[1]:
                cell.fill = PatternFill("solid", fgColor=navy)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for row in rows:
                sheet.append(row)
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column in sheet.columns:
                letter = column[0].column_letter
                width = min(42, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
                sheet.column_dimensions[letter].width = width

        metric_labels = {
            "new_contacts": "Novos contatos", "new_opportunities": "Novas oportunidades",
            "won": "Vendas fechadas", "lost": "Negócios perdidos", "conversion_rate": "Conversão (%)",
            "pipeline_value": "Pipeline aberto (R$)", "won_value": "Volume fechado (R$)",
            "average_cycle_days": "Ciclo médio (dias)", "overdue_tasks": "Tarefas atrasadas",
            "completed_visits": "Visitas concluídas", "sent_proposals": "Propostas movimentadas",
        }
        summary.append(["RELATÓRIO COMERCIAL IN MARE"])
        summary["A1"].font = Font(size=16, bold=True, color=navy)
        summary.append(["Período", str(payload["period"]["date_from"]), str(payload["period"]["date_to"])])
        summary.append([])
        summary.append(["Indicador", "Valor"])
        for cell in summary[4]:
            cell.fill = PatternFill("solid", fgColor=gold)
            cell.font = Font(bold=True, color=navy)
        for key, value in payload["metrics"].items():
            summary.append([metric_labels.get(key, key), value])
        summary.column_dimensions["A"].width = 32
        summary.column_dimensions["B"].width = 20
        summary.column_dimensions["C"].width = 16

        sheet = workbook.create_sheet("Funil")
        write_sheet(sheet, ["Etapa", "Oportunidades", "Valor (R$)"], [[row["label"], row["total"], row["value"]] for row in payload["by_stage"]])
        sheet = workbook.create_sheet("Origens")
        write_sheet(sheet, ["Origem", "Oportunidades"], [[row["source"] or "Não informada", row["total"]] for row in payload["by_source"]])
        sheet = workbook.create_sheet("Equipe")
        write_sheet(sheet, ["Corretor", "Oportunidades", "Ganhas", "Perdidas", "Conversão (%)", "Volume (R$)", "Atrasos"], [
            [row["broker_name"], row["opportunities"], row["won"], row["lost"], row["conversion_rate"], row["won_value"], row["overdue_tasks"]]
            for row in payload["broker_performance"]
        ])
        sheet = workbook.create_sheet("Evolução diária")
        write_sheet(sheet, ["Data", "Novas oportunidades", "Vendas fechadas"], [[row["date"], row["opportunities"], row["won"]] for row in payload["trend"]])
        sheet = workbook.create_sheet("Motivos de perda")
        write_sheet(sheet, ["Motivo", "Quantidade"], [[row["loss_reason"], row["total"]] for row in payload["loss_reasons"]])

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{cls._filename(payload, "xlsx")}"'
        return response

    @classmethod
    def _pdf_response(cls, payload):
        output = BytesIO()
        document = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
        styles = getSampleStyleSheet()
        story = [Paragraph("Relatório comercial In Mare", styles["Title"]), Paragraph(
            f"Período: {payload['period']['date_from']} a {payload['period']['date_to']}", styles["Normal"]
        ), Spacer(1, 5 * mm)]
        metrics = payload["metrics"]
        summary_rows = [
            ["Novos contatos", "Oportunidades", "Conversão", "Vendas", "Volume fechado", "Pipeline", "Atrasos"],
            [metrics["new_contacts"], metrics["new_opportunities"], f'{metrics["conversion_rate"]}%', metrics["won"],
             f'R$ {metrics["won_value"]:,.2f}', f'R$ {metrics["pipeline_value"]:,.2f}', metrics["overdue_tasks"]],
        ]
        header_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B2947")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), .35, colors.HexColor("#C8D2DC")),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6F8")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 7), ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ])
        table = Table(summary_rows, repeatRows=1)
        table.setStyle(header_style)
        story.extend([table, Spacer(1, 6 * mm), Paragraph("Desempenho da equipe", styles["Heading2"])])
        team_rows = [["Corretor", "Oportunidades", "Ganhas", "Perdidas", "Conversão", "Volume", "Atrasos"]]
        team_rows.extend([[row["broker_name"], row["opportunities"], row["won"], row["lost"], f'{row["conversion_rate"]}%', f'R$ {row["won_value"]:,.2f}', row["overdue_tasks"]] for row in payload["broker_performance"]])
        if len(team_rows) == 1:
            team_rows.append(["Sem dados", "-", "-", "-", "-", "-", "-"])
        table = Table(team_rows, repeatRows=1, colWidths=[58 * mm, 28 * mm, 20 * mm, 20 * mm, 25 * mm, 38 * mm, 20 * mm])
        table.setStyle(header_style)
        story.extend([table, Spacer(1, 6 * mm), Paragraph("Funil atual", styles["Heading2"])])
        funnel_rows = [["Etapa", "Oportunidades", "Valor"]] + [[row["label"], row["total"], f'R$ {row["value"]:,.2f}'] for row in payload["by_stage"]]
        table = Table(funnel_rows, repeatRows=1, colWidths=[75 * mm, 40 * mm, 55 * mm])
        table.setStyle(header_style)
        story.append(table)
        document.build(story)
        response = HttpResponse(output.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{cls._filename(payload, "pdf")}"'
        return response


class AdminBrokerViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAdminUser]; serializer_class = BrokerSerializer; queryset = Broker.objects.all()
class AdminUserViewSet(mixins.ListModelMixin, mixins.CreateModelMixin, mixins.UpdateModelMixin, viewsets.GenericViewSet):
    permission_classes = [permissions.IsAdminUser]
    serializer_class = AdminUserSerializer

    def get_queryset(self):
        return get_user_model().objects.filter(is_staff=True).order_by("-is_active", "first_name", "username")

    def perform_create(self, serializer):
        user = serializer.save()
        AuditEvent.objects.create(actor=self.request.user, action="admin.created", entity_type="User", entity_id=str(user.pk), metadata={"username": user.username})

    def perform_update(self, serializer):
        user = serializer.save()
        AuditEvent.objects.create(actor=self.request.user, action="admin.updated", entity_type="User", entity_id=str(user.pk), metadata={"username": user.username, "active": user.is_active})
class AdminSettingsViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAdminUser]; serializer_class = SiteSettingsSerializer; queryset = SiteSettings.objects.all()
    def perform_update(self, serializer):
        old_files = {
            "hero_video": (serializer.instance.hero_video.storage, serializer.instance.hero_video.name)
            if serializer.instance.hero_video else None,
            "hero_poster": (serializer.instance.hero_poster.storage, serializer.instance.hero_poster.name)
            if serializer.instance.hero_poster else None,
        }
        updated = serializer.save()
        for field_name, stored in old_files.items():
            current = getattr(updated, field_name)
            if stored and stored[1] != (current.name if current else ""):
                stored[0].delete(stored[1])
        if self.request.FILES:
            AuditEvent.objects.create(
                actor=self.request.user,
                action="content.hero_video_updated",
                entity_type="SiteSettings",
                entity_id=str(updated.id),
                metadata={"poster_updated": "hero_poster" in self.request.FILES},
            )
    @action(detail=True, methods=["post"], url_path="clear-hero-video")
    def clear_hero_video(self, request, pk=None):
        settings = self.get_object()
        stored_files = [
            (field.storage, field.name)
            for field in (settings.hero_video, settings.hero_poster)
            if field and field.name
        ]
        settings.hero_video = ""
        settings.hero_poster = ""
        settings.hero_video_enabled = False
        settings.save(update_fields=["hero_video", "hero_poster", "hero_video_enabled", "updated_at"])
        for storage, name in stored_files:
            storage.delete(name)
        AuditEvent.objects.create(
            actor=request.user,
            action="content.hero_video_cleared",
            entity_type="SiteSettings",
            entity_id=str(settings.id),
        )
        return Response(self.get_serializer(settings).data)
class AdminHeroSlideViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAdminUser]; serializer_class = HeroSlideSerializer; queryset = HeroSlide.objects.all()
class AdminInstitutionalImageViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAdminUser]; serializer_class = InstitutionalImageSerializer; queryset = InstitutionalImage.objects.all()
class AdminTestimonialViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAdminUser]; serializer_class = TestimonialSerializer; queryset = Testimonial.objects.all()
class AdminFrequentlyAskedQuestionViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAdminUser]; serializer_class = FrequentlyAskedQuestionSerializer; queryset = FrequentlyAskedQuestion.objects.all()
class AdminAuditViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [permissions.IsAdminUser]; serializer_class = AuditSerializer; queryset = AuditEvent.objects.all().select_related("actor").order_by("-created_at")
class AdminImportViewSet(mixins.CreateModelMixin, viewsets.ReadOnlyModelViewSet):
    permission_classes = [permissions.IsAdminUser]; serializer_class = ImportJobSerializer; queryset = ImportJob.objects.all().select_related("property").order_by("-created_at")
    def create(self, request):
        upload = request.FILES.get("file")
        if upload:
            if not upload.name.lower().endswith(".zip"): raise ValidationError({"file": "Envie um arquivo ZIP."})
            with tempfile.NamedTemporaryFile(suffix=".zip") as temporary:
                for chunk in upload.chunks(): temporary.write(chunk)
                temporary.flush()
                job = import_property_zip(temporary.name)
            return Response(self.get_serializer(job).data, status=status.HTTP_201_CREATED)
        path = request.data.get("path")
        if not path: raise ValidationError({"path": "Informe uma pasta disponível no servidor."})
        return Response(self.get_serializer(import_property_folder(path)).data, status=status.HTTP_201_CREATED)

class DashboardView(APIView):
    permission_classes = [permissions.IsAdminUser]
    def get(self, request):
        properties = Property.objects.all()
        return Response({"properties": properties.count(), "published": properties.filter(published=True).count(), "featured": properties.filter(featured=True).count(), "needs_review": sum(p.review_color != "green" for p in properties), "leads": Lead.objects.count(), "imports": ImportJob.objects.count()})
